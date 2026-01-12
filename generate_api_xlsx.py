"""
Script to generate API Documentation XLSX file
Run: pip install openpyxl && python generate_api_xlsx.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
service_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
python_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
signalr_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# All API data
apis = [
    # AccountService
    {"service": "AccountService", "group": "Authentication", "method": "GET", "route": "/account/auth/ok", "input": "None", "output": "string", "description": "Health check endpoint for Account service"},
    {"service": "AccountService", "group": "Authentication", "method": "POST", "route": "/account/auth/login", "input": "Body: LoginRequest {Email: string, Password: string}\nQuery: tenantId (int)", "output": "BaseResponse<LoginResponseDto>", "description": "Authenticate user and return JWT token"},
    {"service": "AccountService", "group": "Account CRUD", "method": "GET", "route": "/account/current-user", "input": "None (from JWT)", "output": "BaseResponse<AccountDTO>", "description": "Get currently authenticated user information"},
    {"service": "AccountService", "group": "Account CRUD", "method": "GET", "route": "/account/{id}", "input": "Route: id (int)", "output": "BaseResponse<AccountDTO>", "description": "Get account details by ID"},
    {"service": "AccountService", "group": "Account CRUD", "method": "GET", "route": "/account/list", "input": "Query: Name (string), Email (string), PageIndex (int), PageSize (int)", "output": "BaseResponse<PaginatedResult<AccountDTO>>", "description": "List accounts with pagination and filtering"},
    {"service": "AccountService", "group": "Account CRUD", "method": "POST", "route": "/account/", "input": "Body: CreateAccountRequest", "output": "BaseResponse<AccountDTO>", "description": "Create a new user account"},
    {"service": "AccountService", "group": "Account CRUD", "method": "POST", "route": "/account/admin-account", "input": "Body: CreateAdminAccountRequest", "output": "BaseResponse<AccountDTO>", "description": "Create a new admin account (restricted)"},
    {"service": "AccountService", "group": "Account CRUD", "method": "PUT", "route": "/account/change-password", "input": "Body: ChangePasswordRequest", "output": "BaseResponse", "description": "Change password for authenticated user"},
    {"service": "AccountService", "group": "Account CRUD", "method": "PUT", "route": "/account/", "input": "Body: UpdateAccountRequest", "output": "BaseResponse<AccountDTO>", "description": "Update account information"},
    {"service": "AccountService", "group": "Account CRUD", "method": "DELETE", "route": "/account/{id}", "input": "Route: id (int)", "output": "BaseResponse", "description": "Delete account by ID"},
    {"service": "AccountService", "group": "Tenant Management", "method": "POST", "route": "/account/tenancy-deactivate", "input": "Query: tenantId (int)", "output": "BaseResponse", "description": "Deactivate tenant (super admin only)"},

    # TenantService
    {"service": "TenantService", "group": "Health Check", "method": "GET", "route": "/tenant/ok", "input": "None", "output": "string", "description": "Health check endpoint for Tenant service"},
    {"service": "TenantService", "group": "Tenant CRUD", "method": "GET", "route": "/tenant/list", "input": "Query: PageIndex (int), PageSize (int)", "output": "BaseResponse<PaginatedResult<TenantDTO>>", "description": "List all tenants with pagination"},
    {"service": "TenantService", "group": "Tenant CRUD", "method": "GET", "route": "/tenant/{id}", "input": "Route: id (int)", "output": "BaseResponse<TenantDTO>", "description": "Get tenant details by ID"},
    {"service": "TenantService", "group": "Tenant CRUD", "method": "POST", "route": "/tenant/create", "input": "Body: CreateTenantRequest {Name, Description, IsActive, Email, Password, AccountName, PermissionsList}", "output": "BaseResponse<TenantDTO>", "description": "Create new tenant with admin account"},
    {"service": "TenantService", "group": "Tenant CRUD", "method": "POST", "route": "/tenant/update", "input": "Body: UpdateTenantRequest", "output": "BaseResponse<TenantDTO>", "description": "Update tenant information"},
    {"service": "TenantService", "group": "Tenant CRUD", "method": "POST", "route": "/tenant/deactivate", "input": "Body: DeactivateTenantRequest", "output": "BaseResponse", "description": "Deactivate a tenant"},

    # ChatService - Conversation
    {"service": "ChatService", "group": "Conversation", "method": "POST", "route": "/chat/conversations", "input": "Body: CreateConversationRequest {Title: string (max 500)}", "output": "BaseResponse<ConversationDTO>", "description": "Create a new conversation"},
    {"service": "ChatService", "group": "Conversation", "method": "GET", "route": "/chat/conversations/list", "input": "None", "output": "BaseResponse<List<ConversationDTO>>", "description": "Get all conversations for current user"},
    {"service": "ChatService", "group": "Conversation", "method": "GET", "route": "/chat/conversations/{conversationId}", "input": "Route: conversationId (int)", "output": "BaseResponse<ConversationDTO>", "description": "Get conversation with all messages"},

    # ChatService - Message
    {"service": "ChatService", "group": "Message", "method": "POST", "route": "/chat/messages", "input": "Body: SendMessageRequest {ConversationId: int, Message: string}", "output": "BaseResponse<MessageDTO>", "description": "Send message and publish to processing queue"},
    {"service": "ChatService", "group": "Message", "method": "GET", "route": "/chat/messages/count", "input": "None", "output": "BaseResponse<int>", "description": "Get total message count"},

    # ChatService - Feedback
    {"service": "ChatService", "group": "Chat Feedback", "method": "GET", "route": "/chat/chat-feedback/", "input": "Query: Ratings (short), PageIndex (int), PageSize (int)", "output": "BaseResponse<PaginatedResult<ChatFeedbackDTO>>", "description": "List chat feedback with pagination"},
    {"service": "ChatService", "group": "Chat Feedback", "method": "GET", "route": "/chat/chat-feedback/{id}", "input": "Route: id (int)", "output": "BaseResponse<ChatFeedbackDTO>", "description": "Get chat feedback by ID"},
    {"service": "ChatService", "group": "Chat Feedback", "method": "POST", "route": "/chat/chat-feedback/", "input": "Body: CreateChatFeedbackRequest", "output": "BaseResponse<ChatFeedbackDTO>", "description": "Create new chat feedback"},
    {"service": "ChatService", "group": "Chat Feedback", "method": "PUT", "route": "/chat/chat-feedback/{id}", "input": "Route: id (int)\nBody: UpdateChatFeedbackRequest", "output": "BaseResponse<ChatFeedbackDTO>", "description": "Update chat feedback"},
    {"service": "ChatService", "group": "Chat Feedback", "method": "POST", "route": "/chat/chat-feedback/rate", "input": "Body: RateChatFeedbackRequest", "output": "BaseResponse", "description": "Submit rating for chat feedback"},

    # ChatService - System Prompt
    {"service": "ChatService", "group": "System Prompt", "method": "GET", "route": "/chat/system-prompt/", "input": "Query: Name (string), IsActive (int: -1/0/1), PageIndex, PageSize", "output": "BaseResponse<PaginatedResult<SystemPromptDTO>>", "description": "List system prompts with filtering"},
    {"service": "ChatService", "group": "System Prompt", "method": "GET", "route": "/chat/system-prompt/{id}", "input": "Route: id (int)", "output": "BaseResponse<SystemPromptDTO>", "description": "Get system prompt by ID"},
    {"service": "ChatService", "group": "System Prompt", "method": "POST", "route": "/chat/system-prompt/", "input": "Body: CreateSystemPromptRequest {Name, Content, Description, IsActive}", "output": "BaseResponse<SystemPromptDTO>", "description": "Create new system prompt"},
    {"service": "ChatService", "group": "System Prompt", "method": "PUT", "route": "/chat/system-prompt/{id}", "input": "Route: id (int)\nBody: UpdateSystemPromptRequest", "output": "BaseResponse<SystemPromptDTO>", "description": "Update system prompt"},
    {"service": "ChatService", "group": "System Prompt", "method": "DELETE", "route": "/chat/system-prompt/{id}", "input": "Route: id (int)", "output": "BaseResponse", "description": "Delete system prompt"},

    # ChatService - Prompt Config
    {"service": "ChatService", "group": "Prompt Config", "method": "GET", "route": "/chat/prompt-config/", "input": "Query: Key (string), PageIndex (int), PageSize (int)", "output": "BaseResponse<PaginatedResult<PromptConfigDTO>>", "description": "List prompt configurations"},
    {"service": "ChatService", "group": "Prompt Config", "method": "GET", "route": "/chat/prompt-config/{id}", "input": "Route: id (int)", "output": "BaseResponse<PromptConfigDTO>", "description": "Get prompt config by ID"},
    {"service": "ChatService", "group": "Prompt Config", "method": "POST", "route": "/chat/prompt-config/", "input": "Body: CreatePromptConfigRequest {Key: string, Value: string}", "output": "BaseResponse<PromptConfigDTO>", "description": "Create new prompt configuration"},
    {"service": "ChatService", "group": "Prompt Config", "method": "PUT", "route": "/chat/prompt-config/{id}", "input": "Route: id (int)\nBody: CreatePromptConfigRequest", "output": "BaseResponse<PromptConfigDTO>", "description": "Update prompt configuration"},
    {"service": "ChatService", "group": "Prompt Config", "method": "DELETE", "route": "/chat/prompt-config/{id}", "input": "Route: id (int)", "output": "BaseResponse", "description": "Delete prompt configuration"},

    # ChatService - SignalR Hub
    {"service": "ChatService", "group": "SignalR Hub", "method": "HUB", "route": "/chatHub -> SendMessage", "input": "conversationId (int), message (string), userId (int)", "output": "Broadcasts: ReceiveMessage", "description": "Send real-time message via SignalR"},
    {"service": "ChatService", "group": "SignalR Hub", "method": "HUB", "route": "/chatHub -> JoinConversation", "input": "conversationId (int)", "output": "Adds to group", "description": "Join conversation group for real-time updates"},
    {"service": "ChatService", "group": "SignalR Hub", "method": "HUB", "route": "/chatHub -> LeaveConversation", "input": "conversationId (int)", "output": "Removes from group", "description": "Leave conversation group"},
    {"service": "ChatService", "group": "SignalR Hub", "method": "HUB", "route": "/chatHub -> OnConnectedAsync", "input": "None", "output": "Connection established", "description": "Client connection lifecycle event"},
    {"service": "ChatService", "group": "SignalR Hub", "method": "HUB", "route": "/chatHub -> OnDisconnectedAsync", "input": "None", "output": "Connection closed", "description": "Client disconnection lifecycle event"},
    {"service": "ChatService", "group": "SignalR Hub", "method": "HUB", "route": "/chatHub -> BroadcastBotResponse", "input": "conversationId (int), messageDto (object)", "output": "Broadcasts: BotResponse", "description": "Broadcast bot response to conversation group"},

    # DocumentService
    {"service": "DocumentService", "group": "Health Check", "method": "GET", "route": "/document/ok", "input": "None", "output": "string", "description": "Health check endpoint for Document service"},
    {"service": "DocumentService", "group": "Document CRUD", "method": "GET", "route": "/document/{id}", "input": "Route: id (int)", "output": "BaseResponse<DocumentDTO>", "description": "Get document details by ID"},
    {"service": "DocumentService", "group": "Document CRUD", "method": "GET", "route": "/document/list", "input": "Query: FileName, UploadedBy, Action (enum), IsApproved (bool), PageIndex, PageSize", "output": "BaseResponse<PaginatedResult<DocumentDTO>>", "description": "List documents with filtering"},
    {"service": "DocumentService", "group": "Document CRUD", "method": "POST", "route": "/document/list-ids", "input": "Body: List<int> (document IDs)", "output": "BaseResponse<List<DocumentDTO>>", "description": "Get multiple documents by IDs"},
    {"service": "DocumentService", "group": "Document CRUD", "method": "POST", "route": "/document/", "input": "Form: file (IFormFile), doctype (enum), fatherDocumentId (int), documentName (string)", "output": "BaseResponse<DocumentDTO>", "description": "Upload and create new document"},
    {"service": "DocumentService", "group": "Document CRUD", "method": "PUT", "route": "/document/", "input": "Body: UpdateDocumentRequest", "output": "BaseResponse<DocumentDTO>", "description": "Update document metadata"},
    {"service": "DocumentService", "group": "Document CRUD", "method": "DELETE", "route": "/document/{id}", "input": "Route: id (int)", "output": "BaseResponse", "description": "Delete document by ID"},
    {"service": "DocumentService", "group": "Vectorization", "method": "POST", "route": "/document/vectorize/{documentId}", "input": "Route: documentId (int)", "output": "BaseResponse", "description": "Vectorize document for semantic search"},

    # StorageService
    {"service": "StorageService", "group": "Local Storage", "method": "POST", "route": "/storage/upload-file", "input": "Form: file (IFormFile), fileName (string), directory (string)", "output": "BaseResponse<FileUploadResponseDTO>", "description": "Upload file to local file system"},
    {"service": "StorageService", "group": "Local Storage", "method": "GET", "route": "/storage/download-file", "input": "Query: filePath (string)", "output": "File stream (octet-stream)", "description": "Download file from local storage"},
    {"service": "StorageService", "group": "MinIO Storage", "method": "POST", "route": "/storage/upload-minio-file", "input": "Form: file (IFormFile), fileName (string), directory (string)", "output": "BaseResponse<FileUploadResponseDTO>", "description": "Upload file to MinIO object storage"},
    {"service": "StorageService", "group": "MinIO Storage", "method": "GET", "route": "/storage/download-minio-file", "input": "Query: filePath (string)", "output": "File stream (octet-stream)", "description": "Download file from MinIO"},

    # ChatProcessor (Python)
    {"service": "ChatProcessor (Python)", "group": "Health Check", "method": "GET", "route": "/health", "input": "None", "output": "{status, ollama: bool, qdrant: bool}", "description": "Check health of Ollama and Qdrant services"},
    {"service": "ChatProcessor (Python)", "group": "Chat Processing", "method": "POST", "route": "/api/chat/test", "input": "Body: ChatRequest {conversation_id, message, user_id, tenant_id, system_instruction}", "output": "ChatResponse {conversation_id, message, timestamp, model_used, rag_documents_used, source_ids, scenario}", "description": "Process chat message with RAG pipeline"},
    {"service": "ChatProcessor (Python)", "group": "Chat Processing", "method": "POST", "route": "/api/test/batch", "input": "Body: BatchTestRequest {entities: [{tenant_id, TC_id, questions}]}", "output": "{status, message, output_file}", "description": "Batch test processing (async, 202 Accepted)"},
    {"service": "ChatProcessor (Python)", "group": "Evaluation", "method": "POST", "route": "/evaluate-batch", "input": "None (uses evaluation_logs.json)", "output": "Evaluation summary", "description": "Run batch evaluation on chat logs"},

    # EmbeddingService (Python)
    {"service": "EmbeddingService (Python)", "group": "Health Check", "method": "GET", "route": "/health", "input": "None", "output": "{status, model, qdrant}", "description": "Check health of EmbeddingService"},
    {"service": "EmbeddingService (Python)", "group": "Embedding", "method": "POST", "route": "/embed", "input": "Body: EmbeddingRequest {text: string}", "output": "EmbeddingResponse {vector: float[], dimensions: int}", "description": "Create embedding vector for text"},
    {"service": "EmbeddingService (Python)", "group": "Vector Storage", "method": "POST", "route": "/vectorize", "input": "Body: VectorizeRequest {text, metadata, collection_name}", "output": "VectorizeResponse {success, point_id, dimensions, collection}", "description": "Create embedding and store in Qdrant"},
    {"service": "EmbeddingService (Python)", "group": "Vector Storage", "method": "POST", "route": "/vectorize-batch", "input": "Body: BatchVectorizeRequest {items: VectorizeRequest[], collection_name}", "output": "VectorizeResponse {success, count, collection}", "description": "Batch vectorize and store documents"},
    {"service": "EmbeddingService (Python)", "group": "Vector Storage", "method": "POST", "route": "/api/embeddings/delete", "input": "Body: DeleteRequest {source_id, tenant_id, type, collection_name}", "output": "VectorizeResponse {success, collection, message}", "description": "Delete vectors by filter criteria"},
    {"service": "EmbeddingService (Python)", "group": "Search", "method": "POST", "route": "/search", "input": "Body: SearchRequest {query, tenant_id, limit, score_threshold}", "output": "List<EmbeddingResponse>", "description": "Search similar documents in Qdrant"},
]

# Create main sheet
ws = wb.active
ws.title = "All APIs"

# Headers
headers = ["STT", "Service", "Group", "Method", "Route", "Input Parameters", "Output/Response", "Description"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Data rows
for idx, api in enumerate(apis, 1):
    row = idx + 1
    ws.cell(row=row, column=1, value=idx).border = thin_border
    ws.cell(row=row, column=2, value=api["service"]).border = thin_border
    ws.cell(row=row, column=3, value=api["group"]).border = thin_border

    method_cell = ws.cell(row=row, column=4, value=api["method"])
    method_cell.border = thin_border
    method_cell.alignment = Alignment(horizontal='center')

    # Color code methods
    if api["method"] == "GET":
        method_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    elif api["method"] == "POST":
        method_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        method_cell.font = Font(color="FFFFFF")
    elif api["method"] == "PUT":
        method_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    elif api["method"] == "DELETE":
        method_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        method_cell.font = Font(color="FFFFFF")
    elif api["method"] == "HUB":
        method_cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        method_cell.font = Font(color="FFFFFF")

    ws.cell(row=row, column=5, value=api["route"]).border = thin_border
    ws.cell(row=row, column=6, value=api["input"]).border = thin_border
    ws.cell(row=row, column=6).alignment = wrap_alignment
    ws.cell(row=row, column=7, value=api["output"]).border = thin_border
    ws.cell(row=row, column=7).alignment = wrap_alignment
    ws.cell(row=row, column=8, value=api["description"]).border = thin_border
    ws.cell(row=row, column=8).alignment = wrap_alignment

    # Highlight Python services
    if "Python" in api["service"]:
        for col in [2, 3]:
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Set column widths
column_widths = [5, 25, 18, 8, 40, 45, 40, 45]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Create summary sheet
ws_summary = wb.create_sheet("Summary")

# Summary headers
summary_headers = ["Service", "Platform", "Total Endpoints", "GET", "POST", "PUT", "DELETE", "HUB"]
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Summary data
summary_data = [
    ["AccountService", ".NET 8", 11, 4, 4, 2, 1, 0],
    ["TenantService", ".NET 8", 6, 2, 3, 0, 0, 0],
    ["ChatService", ".NET 8 + SignalR", 26, 8, 6, 3, 3, 6],
    ["DocumentService", ".NET 8", 8, 3, 3, 1, 1, 0],
    ["StorageService", ".NET 8", 4, 2, 2, 0, 0, 0],
    ["ChatProcessor", "Python FastAPI", 4, 1, 3, 0, 0, 0],
    ["EmbeddingService", "Python FastAPI", 6, 1, 5, 0, 0, 0],
    ["TOTAL", "", 65, 21, 26, 6, 5, 6],
]

for row_idx, row_data in enumerate(summary_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == len(summary_data) + 1:  # Total row
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Set summary column widths
summary_widths = [20, 18, 15, 8, 8, 8, 8, 8]
for col, width in enumerate(summary_widths, 1):
    ws_summary.column_dimensions[get_column_letter(col)].width = width

# Save workbook
output_path = r"G:\repos\AIChat2025\4.2.3_API_Documentation.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
print(f"Total APIs documented: {len(apis)}")
